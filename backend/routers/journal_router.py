# routers/journal_router.py
# ─────────────────────────────────────────────────────────────────────────────
#  Journal des actions correctives
#  GET  /api/journal/{serre_id}              — événements filtrables
#  GET  /api/journal/{serre_id}/resume       — résumé du jour (nb activations)
#  GET  /api/journal/{serre_id}/frequence    — fréquence par jour sur N jours
#  GET  /api/journal/{serre_id}/export       — export CSV ou Excel
# ─────────────────────────────────────────────────────────────────────────────

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from database import get_db
from auth import get_current_user
from datetime import date, datetime, timedelta, timezone
from typing import Optional
import io, csv

router = APIRouter(prefix="/api/journal", tags=["Journal"])

# ── Actions valides ────────────────────────────────────────────────────────
ACTIONS_LABELS = {
    "ventilation":      {"fr": "Ventilation",      "en": "Ventilation"},
    "chauffage":        {"fr": "Chauffage",         "en": "Heating"},
    "brumisateur":      {"fr": "Brumisateur",       "en": "Humidifier"},
    "deshumidification":{"fr": "Déshumidification", "en": "Dehumidification"},
    "co2_injection":    {"fr": "CO₂ Injection",     "en": "CO₂ Injection"},
    "co2_purge":        {"fr": "CO₂ Purge",         "en": "CO₂ Purge"},
}


# ── GET /api/journal/{serre_id} ───────────────────────────────────────────
@router.get("/{serre_id}")
async def get_journal(
    serre_id: int,
    date_debut: Optional[str] = Query(None, description="YYYY-MM-DD"),
    date_fin:   Optional[str] = Query(None, description="YYYY-MM-DD"),
    action:     Optional[str] = Query(None),
    etat:       Optional[str] = Query(None, description="actif | inactif"),
    limit:      int           = Query(200, le=1000),
    db=Depends(get_db),
    user=Depends(get_current_user)
):
    """Liste des événements du journal, filtrables par date, action et état."""

    conditions = ["serre_id = $1"]
    params     = [serre_id]
    i          = 2

    if date_debut:
        conditions.append(f"timestamp >= ${i}::date")
        params.append(date_debut); i += 1
    else:
        # Par défaut : 7 derniers jours
        conditions.append(f"timestamp >= NOW() - INTERVAL '7 days'")

    if date_fin:
        conditions.append(f"timestamp < (${i}::date + INTERVAL '1 day')")
        params.append(date_fin); i += 1

    if action and action in ACTIONS_LABELS:
        conditions.append(f"action = ${i}")
        params.append(action); i += 1

    if etat in ("actif", "inactif"):
        conditions.append(f"etat = ${i}")
        params.append(etat); i += 1

    where = " AND ".join(conditions)
    rows  = await db.fetch(f"""
        SELECT id, serre_id, action, etat, valeur_capteur, seuil_reference, periode, timestamp
        FROM journal_actions
        WHERE {where}
        ORDER BY timestamp DESC
        LIMIT ${i}
    """, *params, limit)

    return [dict(r) for r in rows]


# ── GET /api/journal/{serre_id}/resume ───────────────────────────────────
@router.get("/{serre_id}/resume")
async def get_resume(
    serre_id: int,
    date_cible: Optional[str] = Query(None, description="YYYY-MM-DD, défaut = aujourd'hui"),
    db=Depends(get_db),
    user=Depends(get_current_user)
):
    """
    Résumé d'une journée : nombre d'activations par action,
    avec les timestamps de chaque activation.
    Format : { date, serreId, actions: { ventilation: { count, activations: [...] } } }
    """
    if date_cible:
        try:
            jour = date.fromisoformat(date_cible)
        except ValueError:
            raise HTTPException(status_code=400, detail="Format date invalide (YYYY-MM-DD)")
    else:
        jour = date.today()

    debut = datetime(jour.year, jour.month, jour.day, tzinfo=timezone.utc)
    fin   = debut + timedelta(days=1)

    rows = await db.fetch("""
        SELECT action, etat, valeur_capteur, periode, timestamp
        FROM journal_actions
        WHERE serre_id = $1
          AND timestamp >= $2
          AND timestamp <  $3
        ORDER BY action, timestamp ASC
    """, serre_id, debut, fin)

    result = {}
    for r in rows:
        a = r["action"]
        if a not in result:
            result[a] = {"activations": [], "inactivations": [], "count": 0}
        entry = {
            "timestamp":       r["timestamp"].isoformat(),
            "valeur_capteur":  float(r["valeur_capteur"]) if r["valeur_capteur"] is not None else None,
            "periode":         r["periode"],
        }
        if r["etat"] == "actif":
            result[a]["activations"].append(entry)
            result[a]["count"] += 1
        else:
            result[a]["inactivations"].append(entry)

    return {
        "date":     jour.isoformat(),
        "serre_id": serre_id,
        "actions":  result,
    }


# ── GET /api/journal/{serre_id}/frequence ────────────────────────────────
@router.get("/{serre_id}/frequence")
async def get_frequence(
    serre_id: int,
    jours:  int           = Query(7, ge=1, le=90, description="Nombre de jours"),
    action: Optional[str] = Query(None),
    db=Depends(get_db),
    user=Depends(get_current_user)
):
    """
    Fréquence d'activation par action, agrégée par jour sur N jours.
    Utile pour les graphiques bar chart.
    """
    conditions = ["serre_id = $1", "etat = 'actif'",
                  f"timestamp >= NOW() - ($2 * INTERVAL '1 day')"]
    params     = [serre_id, jours]
    i          = 3

    if action and action in ACTIONS_LABELS:
        conditions.append(f"action = ${i}")
        params.append(action); i += 1

    where = " AND ".join(conditions)
    rows  = await db.fetch(f"""
        SELECT
            DATE(timestamp AT TIME ZONE 'Africa/Casablanca') AS jour,
            action,
            COUNT(*) AS nb_activations
        FROM journal_actions
        WHERE {where}
        GROUP BY jour, action
        ORDER BY jour ASC, action ASC
    """, *params)

    return [{"jour": str(r["jour"]), "action": r["action"], "count": int(r["nb_activations"])} for r in rows]


# ── GET /api/journal/{serre_id}/export ───────────────────────────────────
@router.get("/{serre_id}/export")
async def export_journal(
    serre_id:   int,
    format:     str           = Query("csv", description="csv | excel"),
    date_debut: Optional[str] = Query(None),
    date_fin:   Optional[str] = Query(None),
    action:     Optional[str] = Query(None),
    db=Depends(get_db),
    user=Depends(get_current_user)
):
    """Export CSV ou Excel du journal pour une serre."""

    serre = await db.fetchrow("SELECT code, nom_fr FROM serres WHERE id = $1", serre_id)
    if not serre:
        raise HTTPException(status_code=404, detail="Serre introuvable")

    conditions = ["j.serre_id = $1"]
    params     = [serre_id]
    i          = 2

    if date_debut:
        conditions.append(f"j.timestamp >= ${i}::date")
        params.append(date_debut); i += 1
    else:
        conditions.append("j.timestamp >= NOW() - INTERVAL '30 days'")

    if date_fin:
        conditions.append(f"j.timestamp < (${i}::date + INTERVAL '1 day')")
        params.append(date_fin); i += 1

    if action and action in ACTIONS_LABELS:
        conditions.append(f"j.action = ${i}")
        params.append(action); i += 1

    where = " AND ".join(conditions)
    rows  = await db.fetch(f"""
        SELECT
            j.timestamp AT TIME ZONE 'Africa/Casablanca' AS ts_local,
            j.action,
            j.etat,
            j.valeur_capteur,
            j.seuil_reference,
            j.periode
        FROM journal_actions j
        WHERE {where}
        ORDER BY j.timestamp ASC
    """, *params)

    HEADERS = ["Date/Heure (Rabat)", "Action", "État", "Valeur capteur", "Seuil référence", "Période"]
    nom_fichier = f"SDI_Journal_{serre['code']}"
    if date_debut: nom_fichier += f"_{date_debut}"
    if date_fin:   nom_fichier += f"_au_{date_fin}"

    ACTION_LABELS_FR = {
        "ventilation":      "Ventilation",
        "chauffage":        "Chauffage",
        "brumisateur":      "Brumisateur",
        "deshumidification":"Déshumidification",
        "co2_injection":    "CO₂ Injection",
        "co2_purge":        "CO₂ Purge",
    }

    def to_row(r):
        return [
            r["ts_local"].strftime("%Y-%m-%d %H:%M") if r["ts_local"] else "",
            ACTION_LABELS_FR.get(r["action"], r["action"]),
            "Actif" if r["etat"] == "actif" else "Inactif",
            str(r["valeur_capteur"]) if r["valeur_capteur"] is not None else "",
            str(r["seuil_reference"]) if r["seuil_reference"] is not None else "",
            "Jour" if r["periode"] == "jour" else "Nuit" if r["periode"] else "",
        ]

    # ── CSV ──────────────────────────────────────────────────
    if format == "csv":
        buf = io.StringIO()
        w   = csv.writer(buf)
        w.writerow(HEADERS)
        for r in rows:
            w.writerow(to_row(r))
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv; charset=utf-8-sig",
            headers={"Content-Disposition": f'attachment; filename="{nom_fichier}.csv"'}
        )

    # ── Excel ────────────────────────────────────────────────
    if format == "excel":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl non installé")

        wb  = Workbook()
        ws  = wb.active
        ws.title = "Journal Actions"

        # Titre
        ws.merge_cells("A1:F1")
        ws["A1"] = f"Journal des actions correctives — {serre['nom_fr']} ({serre['code']})"
        ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
        ws["A1"].fill      = PatternFill("solid", fgColor="0F172A")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        # En-têtes
        for ci, h in enumerate(HEADERS, 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.font      = Font(bold=True, color="FFFFFF", size=11)
            c.fill      = PatternFill("solid", fgColor="1B4332")
            c.alignment = Alignment(horizontal="center")

        ws.freeze_panes = "A3"

        # Couleurs par état
        fill_actif   = PatternFill("solid", fgColor="D1FAE5")
        fill_inactif = PatternFill("solid", fgColor="F1F5F9")
        alt_fill     = PatternFill("solid", fgColor="F8FAFC")
        center       = Alignment(horizontal="center")

        for ri, r in enumerate(rows):
            row_data = to_row(r)
            ws.append(row_data)
            rn = ri + 3
            is_actif = r["etat"] == "actif"
            for ci in range(1, 7):
                cell = ws.cell(row=rn, column=ci)
                cell.alignment = center
                if ci == 3:  # colonne État
                    cell.fill = fill_actif if is_actif else fill_inactif
                    cell.font = Font(bold=True, color="166534" if is_actif else "64748B")
                elif ri % 2 == 1:
                    cell.fill = alt_fill

        for i_col, width in enumerate([20, 20, 10, 16, 18, 10], 1):
            ws.column_dimensions[get_column_letter(i_col)].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{nom_fichier}.xlsx"'}
        )

    raise HTTPException(status_code=400, detail="Format invalide. Utilisez 'csv' ou 'excel'.")
