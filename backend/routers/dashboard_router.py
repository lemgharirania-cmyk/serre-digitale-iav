# routers/dashboard_router.py — Endpoints privés gérants
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
from database import get_db
from auth import get_current_user
import io
import csv

router = APIRouter(prefix="/api/dashboard", tags=["Dashboard"])

# ─── Thresholds ─────────────────────────────────────────────

class ThresholdUpdate(BaseModel):
    valeur_min:   Optional[float] = None
    valeur_max:   Optional[float] = None
    email_alerte: Optional[str]   = None
    actif:        Optional[bool]  = True

@router.get("/thresholds")
async def get_all_thresholds(db=Depends(get_db), user=Depends(get_current_user)):
    rows = await db.fetch("""
        SELECT t.*, s.nom_fr, s.code
        FROM thresholds t
        JOIN serres s ON s.id = t.serre_id
        ORDER BY s.code, t.capteur
    """)
    return [dict(r) for r in rows]

@router.get("/thresholds/{serre_id}")
async def get_serre_thresholds(serre_id: int, db=Depends(get_db), user=Depends(get_current_user)):
    rows = await db.fetch(
        "SELECT * FROM thresholds WHERE serre_id=$1 ORDER BY capteur", serre_id
    )
    return [dict(r) for r in rows]

@router.put("/thresholds/{serre_id}/{capteur}")
async def update_threshold(
    serre_id: int, capteur: str,
    data: ThresholdUpdate,
    db=Depends(get_db),
    user=Depends(get_current_user)
):
    await db.execute("""
        INSERT INTO thresholds (serre_id, capteur, valeur_min, valeur_max, email_alerte, actif, updated_at)
        VALUES ($1, $2, $3, $4, $5, $6, NOW())
        ON CONFLICT (serre_id, capteur)
        DO UPDATE SET
            valeur_min   = EXCLUDED.valeur_min,
            valeur_max   = EXCLUDED.valeur_max,
            email_alerte = EXCLUDED.email_alerte,
            actif        = EXCLUDED.actif,
            updated_at   = NOW()
    """, serre_id, capteur, data.valeur_min, data.valeur_max, data.email_alerte, data.actif)
    return {"message": f"Seuil {capteur} mis à jour pour serre {serre_id}"}

# ─── Alertes ────────────────────────────────────────────────

@router.get("/alertes")
async def get_alertes(non_lues: bool = False, db=Depends(get_db), user=Depends(get_current_user)):
    query = """
        SELECT a.*, s.nom_fr, s.code
        FROM alertes a JOIN serres s ON s.id = a.serre_id
    """
    if non_lues:
        query += " WHERE a.lu = FALSE"
    query += " ORDER BY a.created_at DESC LIMIT 100"
    rows = await db.fetch(query)
    return [dict(r) for r in rows]

@router.put("/alertes/{alerte_id}/lue")
async def marquer_lue(alerte_id: int, db=Depends(get_db), user=Depends(get_current_user)):
    await db.execute("UPDATE alertes SET lu=TRUE WHERE id=$1", alerte_id)
    return {"message": "Alerte marquée comme lue"}

@router.put("/alertes/tout-lire")
async def tout_lire(db=Depends(get_db), user=Depends(get_current_user)):
    await db.execute("UPDATE alertes SET lu=TRUE WHERE lu=FALSE")
    return {"message": "Toutes les alertes marquées comme lues"}

# ─── Comparaison serres ──────────────────────────────────────

@router.get("/comparaison")
async def comparer_serres(
    capteur: str = "temperature",
    heures: int = 24,
    db=Depends(get_db),
    user=Depends(get_current_user)
):
    CAPTEURS_VALIDES = {
        "temperature", "humidite", "vpd", "co2", "luminosite",
        "ph", "ec", "temp_eau", "niveau_eau"
    }
    if capteur not in CAPTEURS_VALIDES:
        raise HTTPException(status_code=400, detail=f"Capteur invalide: {capteur}")
    serres = await db.fetch("SELECT id, code, nom_fr FROM serres WHERE actif=TRUE")
    result = []
    for serre in serres:
        rows = await db.fetch(f"""
            SELECT capture_at, {capteur} as valeur
            FROM mesures_iot
            WHERE serre_id=$1
              AND {capteur} IS NOT NULL
              AND capture_at > NOW() - ($2 * INTERVAL '1 hour')
            ORDER BY capture_at ASC
        """, serre["id"], heures)
        result.append({
            "serre_id": serre["id"],
            "code":     serre["code"],
            "nom_fr":   serre["nom_fr"],
            "data": [{"time": r["capture_at"].isoformat(), "value": r["valeur"]} for r in rows]
        })
    return result

# ─── Export ─────────────────────────────────────────────────

@router.get("/export/{serre_id}")
async def export_data(
    serre_id: int,
    format: str = "csv",
    heures: int = 168,
    db=Depends(get_db),
    user=Depends(get_current_user)
):
    import datetime as dt

    serre = await db.fetchrow("SELECT * FROM serres WHERE id=$1", serre_id)
    if not serre:
        raise HTTPException(status_code=404, detail="Serre introuvable")

    rows = await db.fetch("""
        SELECT capture_at, type_api,
               temperature, humidite, vpd, co2, luminosite,
               ph, ec, temp_eau, niveau_eau
        FROM mesures_iot
        WHERE serre_id=$1
          AND capture_at > NOW() - ($2 * INTERVAL '1 hour')
        ORDER BY capture_at ASC
    """, serre_id, heures)

    if not rows:
        raise HTTPException(status_code=404, detail="Aucune donnée disponible pour cette période.")

    nom_fichier = f"SDI_{serre['code']}_{heures}h"

    def fmt(v):
        return "" if v is None else v

    # ── Merge ENV + IRR rows by nearest timestamp (30s window) ──
    env_rows = [dict(r) for r in rows if r["type_api"] == "ENV"]
    irr_rows = [dict(r) for r in rows if r["type_api"] == "IRR"]

    merged = []
    used_irr = set()
    for env in env_rows:
        best_irr   = None
        best_delta = dt.timedelta(seconds=30)
        for idx, irr in enumerate(irr_rows):
            if idx in used_irr:
                continue
            delta = abs(env["capture_at"] - irr["capture_at"])
            if delta < best_delta:
                best_delta = delta
                best_irr   = (idx, irr)
        row_m = {
            "capture_at":  env["capture_at"],
            "temperature": env.get("temperature"),
            "humidite":    env.get("humidite"),
            "vpd":         env.get("vpd"),
            "co2":         env.get("co2"),
            "luminosite":  env.get("luminosite"),
            "ph": None, "ec": None, "temp_eau": None, "niveau_eau": None,
        }
        if best_irr:
            used_irr.add(best_irr[0])
            row_m["ph"]         = best_irr[1].get("ph")
            row_m["ec"]         = best_irr[1].get("ec")
            row_m["temp_eau"]   = best_irr[1].get("temp_eau")
            row_m["niveau_eau"] = best_irr[1].get("niveau_eau")
        merged.append(row_m)

    for idx, irr in enumerate(irr_rows):
        if idx not in used_irr:
            merged.append({
                "capture_at": irr["capture_at"],
                "temperature": None, "humidite": None, "vpd": None,
                "co2": None, "luminosite": None,
                "ph": irr.get("ph"), "ec": irr.get("ec"),
                "temp_eau": irr.get("temp_eau"), "niveau_eau": irr.get("niveau_eau"),
            })

    merged.sort(key=lambda r: r["capture_at"])

    HEADERS = [
        "Date/Heure",
        "Température (°C)", "Humidité (%)", "VPD (kPa)", "CO2 (PPM)", "Luminosité (lux)",
        "pH", "EC (mS/cm)", "Temp. Eau (°C)", "Niveau Eau (m)",
    ]

    def to_list(r):
        return [
            r["capture_at"].strftime("%Y-%m-%d %H:%M:%S") if r["capture_at"] else "",
            fmt(r["temperature"]), fmt(r["humidite"]), fmt(r["vpd"]),
            fmt(r["co2"]),         fmt(r["luminosite"]),
            fmt(r["ph"]),          fmt(r["ec"]),
            fmt(r["temp_eau"]),    fmt(r["niveau_eau"]),
        ]

    # ── CSV ──
    if format == "csv":
        buf    = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(HEADERS)
        for r in merged:
            writer.writerow(to_list(r))
        csv_bytes = buf.getvalue().encode("utf-8-sig")
        return StreamingResponse(
            io.BytesIO(csv_bytes),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{nom_fichier}.csv"'}
        )

    # ── EXCEL ──
    if format == "excel":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{serre['code']} {heures}h"

        center   = Alignment(horizontal="center", vertical="center")
        alt_fill = PatternFill("solid", fgColor="F9FAFB")

        # Row 1 — Group labels
        ws.append(["", "Environnement", "", "", "", "", "Irrigation", "", "", ""])
        ws.merge_cells("B1:F1")
        ws.merge_cells("G1:J1")
        ws.row_dimensions[1].height = 18
        for col in range(1, 11):
            c = ws.cell(row=1, column=col)
            c.alignment = center
            if col == 1:
                c.fill = PatternFill("solid", fgColor="F3F4F6")
            elif col <= 6:
                c.fill = PatternFill("solid", fgColor="D1FAE5")
                c.font = Font(bold=True, color="065F46", size=10)
            else:
                c.fill = PatternFill("solid", fgColor="DBEAFE")
                c.font = Font(bold=True, color="1E40AF", size=10)

        # Row 2 — Column headers
        ws.append(HEADERS)
        ws.row_dimensions[2].height = 22
        for col, c in enumerate(ws[2], 1):
            c.alignment = center
            if col == 1:
                c.font = Font(bold=True, color="FFFFFF", size=11)
                c.fill = PatternFill("solid", fgColor="374151")
            elif col <= 6:
                c.font = Font(bold=True, color="FFFFFF", size=11)
                c.fill = PatternFill("solid", fgColor="1B4332")
            else:
                c.font = Font(bold=True, color="FFFFFF", size=11)
                c.fill = PatternFill("solid", fgColor="1E3A5F")

        ws.freeze_panes = "A3"

        # Data rows
        for ri, r in enumerate(merged):
            ws.append(to_list(r))
            rn = ri + 3
            for col in range(1, 11):
                c = ws.cell(row=rn, column=col)
                c.alignment = center
                if ri % 2 == 1:
                    c.fill = alt_fill

        # Column widths
        for i, w in enumerate([18,14,12,10,10,14,8,12,14,13], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{nom_fichier}.xlsx"'}
        )

    raise HTTPException(status_code=400, detail="Format non supporté. Utilisez 'csv' ou 'excel'.")

# ─── Utilisateurs (admin only) ───────────────────────────────

class NouvelUtilisateur(BaseModel):
    nom:          str
    email:        str
    mot_de_passe: str
    role:         str = "gerant"
    serre_id:     Optional[int] = None

@router.get("/utilisateurs")
async def list_users(db=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin requis")
    rows = await db.fetch("SELECT id, nom, email, role, actif, created_at, last_login FROM utilisateurs")
    return [dict(r) for r in rows]

@router.post("/utilisateurs")
async def create_user(data: NouvelUtilisateur, db=Depends(get_db), user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin requis")
    from auth import hash_password
    hashed = hash_password(data.mot_de_passe)
    await db.execute("""
        INSERT INTO utilisateurs (nom, email, mot_de_passe, role)
        VALUES ($1, $2, $3, $4)
    """, data.nom, data.email, hashed, data.role)
    return {"message": f"Utilisateur {data.email} créé"}
